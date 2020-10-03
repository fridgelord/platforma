import pandas
import shutil
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path


class DataFrameAppend(pandas.DataFrame):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)


    def append_to_excel(self, filename, sheet_name='Sheet1', startrow=None,
                           truncate_sheet=False,
                           **to_excel_kwargs):
        """Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        Parameters:
          filename : File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
          df : dataframe to save to workbook
          sheet_name : Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
          startrow : upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
          truncate_sheet : truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
          to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]

        Returns: None
        """

        try:
            # ignore [engine] parameter if it was passed
            if 'engine' in to_excel_kwargs:
                to_excel_kwargs.pop('engine')
            if 'header' in to_excel_kwargs:
                header = to_excel_kwargs['header']
                to_excel_kwargs.pop('header')
            else:
                header = True

            with pandas.ExcelWriter(filename, engine='openpyxl') as writer:
                try:
                    writer.book = load_workbook(filename)

                    # get the last row in the existing Excel sheet
                    # if it was not specified explicitly
                    if startrow is None and sheet_name in writer.book.sheetnames:
                        startrow = writer.book[sheet_name].max_row

                    # truncate sheet
                    if truncate_sheet and sheet_name in writer.book.sheetnames:
                        # index of [sheet_name] sheet
                        idx = writer.book.sheetnames.index(sheet_name)
                        # remove [sheet_name]
                        writer.book.remove(writer.book.worksheets[idx])
                        # create an empty sheet [sheet_name] using old index
                        writer.book.create_sheet(sheet_name, idx)

                    # copy existing sheets
                    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
                except FileNotFoundError:
                    # file does not exist yet, we will create it
                    pass

                if startrow is None:
                    startrow = 0
                else:
                    header = False

                self.to_excel(writer, sheet_name, startrow=startrow, header=header, **to_excel_kwargs)
        except PermissionError:
            old_file = Path(filename)
            root = old_file.root
            stem = old_file.stem
            suffix = old_file.suffix
            new_filename = stem + "_" + datetime.now().strftime("%Y%m%d_%H%M%S") + suffix
            new_file = Path(root).joinpath(new_filename)
            new_filepath = str(new_file)
            shutil.copy(filename, new_filepath)

            self.append_to_excel(new_filepath, sheet_name, startrow,
                       truncate_sheet, **to_excel_kwargs)
